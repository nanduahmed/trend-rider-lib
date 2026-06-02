"""
Helpers for formatting report workbooks and debug CSV output.
"""
from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Optional, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..core.enums import SignalType
from ..core.models import SignalEvent


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBTLE_FILL = PatternFill("solid", fgColor="F3F4F6")
GREEN_FILL = PatternFill("solid", fgColor="E2F5E6")
RED_FILL = PatternFill("solid", fgColor="FDE2E1")
AMBER_FILL = PatternFill("solid", fgColor="FFF5D6")
BLUE_FILL = PatternFill("solid", fgColor="DCEBFF")
NEUTRAL_FILL = PatternFill("solid", fgColor="F9FAFB")

SIGNAL_REASON_MAP: Dict[str, str] = {
    "UPTREND_START": "Weekly close confirmed the official trend start",
    "BUY_ENTRY": "First qualifying bullish crossover after trend qualification",
    "REENTRY": "TR-qualified stock re-entered the buy zone",
    "MOMENTUM_ENTRY": "Qualified bullish crossover during recovery",
    "DOWNTREND_START": "Weekly close fell below the 0.90 * EMA21 trigger",
    "EMA_CROSSOVER": "EMA34 crossed above EMA55",
    "TR_QUALIFIED": "Uptrend weeks reached the qualification threshold",
}

SIGNAL_ROW_FILL_MAP: Dict[str, PatternFill] = {
    "UPTREND_START": GREEN_FILL,
    "BUY_ENTRY": GREEN_FILL,
    "REENTRY": BLUE_FILL,
    "MOMENTUM_ENTRY": BLUE_FILL,
    "DOWNTREND_START": RED_FILL,
    "EMA_CROSSOVER": AMBER_FILL,
    "TR_QUALIFIED": GREEN_FILL,
}

IMPORTANT_SIGNAL_HINTS = (
    "UPTREND_START",
    "BUY_ENTRY",
    "REENTRY",
    "MOMENTUM_ENTRY",
    "DOWNTREND_START",
    "EMA_CROSSOVER",
    "TR_QUALIFIED",
)


@dataclass(frozen=True)
class SheetFormat:
    """Formatting instructions for a single Excel worksheet."""

    date_columns: Sequence[str] = ()
    highlight_column: Optional[str] = None
    highlight_contains: bool = False
    highlight_values: Mapping[str, PatternFill] = field(default_factory=dict)
    freeze_panes: str = "A2"


def safe_filename_part(value: str) -> str:
    """Convert a ticker or label into a safe filename fragment."""
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    cleaned = cleaned.strip("._-")
    return cleaned or "item"


def signal_reason(signal_type: SignalType | str) -> str:
    """Human-readable reason for a signal type."""
    signal_name = signal_type.name if isinstance(signal_type, SignalType) else str(signal_type)
    return SIGNAL_REASON_MAP.get(signal_name, signal_name)


def _derive_candle_color(signal: SignalEvent) -> str:
    metadata = signal.metadata or {}
    open_price = metadata.get("open")
    close_price = metadata.get("close", signal.close_price)
    if open_price is not None and close_price is not None:
        if close_price > open_price:
            return "GREEN"
        if close_price < open_price:
            return "RED"
        return "NEUTRAL"

    if signal.signal_type in {
        SignalType.UPTREND_START,
        SignalType.BUY_ENTRY,
        SignalType.REENTRY,
        SignalType.MOMENTUM_ENTRY,
        SignalType.TR_QUALIFIED,
    }:
        return "GREEN"
    if signal.signal_type == SignalType.DOWNTREND_START:
        return "RED"
    return "YELLOW"


def _build_supporting_values(signal: SignalEvent) -> str:
    metadata = signal.metadata or {}
    base_keys = {
        "reason",
        "timeframe",
        "state",
        "open",
        "high",
        "low",
        "close",
        "volume",
        "ema21",
        "ema34",
        "ema55",
    }

    values = []
    for key in (
        "candle_count",
        "weekly_candle_count",
        "uptrend_weeks",
        "tr_qualified",
        "is_buyzone",
        "is_crossover_detected",
        "trend_cycle_id",
    ):
        if key in metadata:
            values.append(f"{key}={metadata[key]}")

    extra_keys = sorted(key for key in metadata if key not in base_keys)
    for key in extra_keys:
        values.append(f"{key}={metadata[key]}")

    return "; ".join(values)


def build_signal_rows(signals: Sequence[SignalEvent]) -> pd.DataFrame:
    """Build a report-ready DataFrame from persisted signal events."""
    rows = []
    for signal in signals:
        metadata = signal.metadata or {}
        rows.append(
            {
                "Date": signal.date,
                "Ticker": signal.ticker,
                "Signal Type": signal.signal_type.name if isinstance(signal.signal_type, SignalType) else str(signal.signal_type),
                "Signal Reason": metadata.get("reason", signal_reason(signal.signal_type)),
                "Candle Color": _derive_candle_color(signal),
                "Timeframe": signal.timeframe or metadata.get("timeframe", ""),
                "State": signal.state or metadata.get("state", ""),
                "Close": signal.close_price,
                "EMA21": signal.ema21,
                "EMA34": signal.ema34,
                "EMA55": signal.ema55,
                "Trend Cycle": signal.trend_cycle_id,
                "Trend Start": signal.trend_start_date,
                "Trend End": signal.trend_end_date,
                "Supporting Values": _build_supporting_values(signal),
                "Metadata": json.dumps(metadata, default=str) if metadata else "",
            }
        )

    if not rows:
        return pd.DataFrame(
            columns=[
                "Date",
                "Ticker",
                "Signal Type",
                "Signal Reason",
                "Candle Color",
                "Timeframe",
                "State",
                "Close",
                "EMA21",
                "EMA34",
                "EMA55",
                "Trend Cycle",
                "Trend Start",
                "Trend End",
                "Supporting Values",
                "Metadata",
            ]
        )

    return pd.DataFrame(rows)


def format_date_column(series: pd.Series) -> pd.Series:
    """
    Convert a series of datetime values to fixed 'YYYY-MM-DD' string format.
    Handles timezone-aware, timezone-naive, and mixed datetime objects safely
    without raising Excel timezone errors.
    """
    # Convert to datetime and coerce errors to NaT
    dt_series = pd.to_datetime(series, errors="coerce")
    # Strip timezone if present
    if dt_series.dt.tz is not None:
        dt_series = dt_series.dt.tz_localize(None)
    # Format as YYYY-MM-DD string — avoids any Excel timezone issues
    formatted = dt_series.dt.strftime("%Y-%m-%d")
    # Replace NaT / NaNs with None so Excel sees empty cells
    return formatted.where(formatted.notna(), None)


def prepare_debug_csv_frame(
    df: pd.DataFrame,
    date_columns: Sequence[str] = (
        "Date",
        "Crossover Date",
        "Trend Start",
        "Trend End",
        "Daily EMA21 Cross",
        "Daily Downtrend Trigger",
        "First Buy Zone",
        "Positive Crossover",
    ),
) -> pd.DataFrame:
    """Prepare a debug frame for CSV output with ISO dates and 2-decimal numbers."""
    prepared = df.copy()

    for column in date_columns:
        if column in prepared.columns:
            prepared[column] = pd.to_datetime(prepared[column], errors="coerce").dt.strftime("%Y-%m-%d")

    numeric_columns = prepared.select_dtypes(include=["number"]).columns
    for column in numeric_columns:
        if pd.api.types.is_bool_dtype(prepared[column]):
            continue
        prepared[column] = pd.to_numeric(prepared[column], errors="coerce").astype(float).round(2)

    return prepared


def write_debug_csv(output_path: Path, df: pd.DataFrame) -> Path:
    """Write the raw analysis DataFrame to CSV with stable formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    prepared = prepare_debug_csv_frame(df)
    prepared.to_csv(output_path, index=False)
    return output_path


def load_debug_csv_frames(cache_dir: Path, tickers: Sequence[str]) -> pd.DataFrame:
    """Load cached debug CSV files for the given tickers and concatenate them."""
    frames = []
    for ticker in tickers:
        debug_path = cache_dir / f"{safe_filename_part(ticker)}_analysis_debug.csv"
        if not debug_path.exists():
            continue
        frame = pd.read_csv(debug_path)
        frames.append(frame)

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


def _column_fill_for_value(value: Any, mapping: Mapping[str, PatternFill], contains: bool) -> Optional[PatternFill]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    value_str = str(value)
    if contains:
        for key, fill in mapping.items():
            if key in value_str:
                return fill
        return None

    return mapping.get(value_str)


def _auto_width(value_series: pd.Series, header: str) -> float:
    sample_values = value_series.dropna().astype(str).head(100).tolist()
    max_len = max([len(header), *[len(value) for value in sample_values]] or [len(header)])
    return min(max_len + 2, 48)


def _coerce_timezone_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove timezone info from ALL datetime-like columns in the DataFrame.

    Scans every column and converts any tz-aware datetime/timestamp to a
    naive 'YYYY-MM-DD' string so Excel never sees a tz-aware value.
    """
    result = df.copy()
    for col in result.columns:
        # Check if the column dtype is datetime-like or contains datetime objects
        if pd.api.types.is_datetime64_any_dtype(result[col]):
            # Convert to naive datetime string — bypasses Excel tz validation
            dt_series = pd.to_datetime(result[col], errors="coerce")
            if dt_series.dt.tz is not None:
                dt_series = dt_series.dt.tz_localize(None)
            result[col] = dt_series.dt.strftime("%Y-%m-%d")
        elif result[col].dtype == object:
            # Object columns may contain individual datetime objects
            # Check first non-null value to decide
            sample = result[col].dropna()
            if not sample.empty and hasattr(sample.iloc[0], "tzinfo"):
                tz_aware = sample.apply(lambda v: getattr(v, "tzinfo", None) is not None if hasattr(v, "tzinfo") else False)
                if tz_aware.any():
                    result[col] = format_date_column(result[col])
    return result


def write_excel_workbook(
    output_path: Path,
    sheets: Mapping[str, pd.DataFrame],
    formats: Optional[Mapping[str, SheetFormat]] = None,
) -> Path:
    """Write a workbook and apply consistent report formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Pre-process frames: strip timezone from ALL datetime-like columns
    processed_sheets = {}
    formats = formats or {}
    for name, frame in sheets.items():
        processed_sheets[name] = _coerce_timezone_columns(frame)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in processed_sheets.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    formats = formats or {}

    for sheet_name, frame in sheets.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        config = formats.get(sheet_name, SheetFormat())

        if ws.max_row == 0 or ws.max_column == 0:
            continue

        # Header styling
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Date and numeric formatting
        for column_index, column_name in enumerate(frame.columns, start=1):
            column_letter = get_column_letter(column_index)
            series = frame[column_name]

            if column_name in config.date_columns:
                for row_index in range(2, ws.max_row + 1):
                    ws[f"{column_letter}{row_index}"].number_format = "yyyy-mm-dd"
            elif pd.api.types.is_numeric_dtype(series):
                for row_index in range(2, ws.max_row + 1):
                    ws[f"{column_letter}{row_index}"].number_format = "0.00"

            ws.column_dimensions[column_letter].width = _auto_width(series, column_name)

        # Row highlighting for important signals / candles.
        if config.highlight_column and config.highlight_column in frame.columns:
            highlight_column_index = frame.columns.get_loc(config.highlight_column) + 1
            highlight_letter = get_column_letter(highlight_column_index)
            for row_index in range(2, ws.max_row + 1):
                value = ws[f"{highlight_letter}{row_index}"].value
                fill = _column_fill_for_value(value, config.highlight_values, config.highlight_contains)
                if fill is None:
                    continue
                for cell in ws[row_index]:
                    cell.fill = fill

        ws.freeze_panes = config.freeze_panes
        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    return output_path
